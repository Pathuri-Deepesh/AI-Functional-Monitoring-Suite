"""Generate project-tracker.xlsx grouped by day, in plain English.

Output: one sheet "Daily Log". Each day is a merged amber header row with
date + count + a 2-line summary, followed by that day's individual tasks
written in everyday language (a few common tech words kept: API, database,
frontend, backend, button, URL, Slack, Postman, JSON, file).
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / "project-tracker.xlsx"

# ---- Day-level summaries (2 short lines each, kept under ~200 chars total) ----
DAY_SUMMARIES = {
    "2026-04-29": (
        "Got the project off the ground.",
        "Set up the backend server + frontend web page, added a 'submit a URL' feature, and grouped results by health (2xx/3xx/4xx/5xx).",
    ),
    "2026-04-30": (
        "Made the web page actually show stuff.",
        "Built colored count cards on the dashboard, made it auto-refresh every 3 seconds, and saved all data to a file so nothing is lost on restart.",
    ),
    "2026-05-05": (
        "Turned the prototype into something a team could actually use.",
        "Added multiple projects, an API-key vault, custom check timings, detailed speed tracking per URL, and Slack alerts when something breaks.",
    ),
    "2026-05-06": (
        "Demo day — showed the team how the auth feature works.",
        "Walked through API-key login and Basic Auth using public test URLs (httpbin) to prove the headers really go through.",
    ),
    "2026-05-07": (
        "Spent the day making the app look and feel polished.",
        "Replaced ugly browser popups with proper dialog boxes, added toast notifications, smoothed out all the animations, and made the sidebar nicer.",
    ),
    "2026-05-08": (
        "Big day — moved off a plain file and into a real database.",
        "Switched to SQLite, started saving every check forever (with cleanup), built history charts, and shipped a 'Check All' button that emails a full report to Slack.",
    ),
    "2026-05-11": (
        "Made the tool work like Postman.",
        "Added POST/PUT/PATCH, a body editor, custom headers, query params, and assertions. Also added search, filter chips, and pagination so finding URLs is easy.",
    ),
    "2026-05-12": (
        "Added a time-range picker for the history charts + tons of tiny polish.",
        "Picker for 24h / 7d / 30d / 1y / Custom (Datadog-style), plus skeleton loaders, spinners, animated pulses on failing items, smooth transitions, and accessibility fixes.",
    ),
    "2026-05-19": (
        "Massive day — built two huge features end-to-end.",
        "1) Flows: chain many API calls together with variables, retries, smart caching, Slack alerts, dashboards. 2) Prereqs: auto-run login/setup before every check.",
    ),
    "2026-05-20": (
        "Made retries visible in real-time.",
        "When a step is retrying or waiting before the next try, the user can now see exactly what's happening — no more staring at a frozen spinner.",
    ),
    "2026-05-21": (
        "Added file uploads + polished the editor to match Postman exactly.",
        "URLs can now send images, PDFs, any binary file. Built an upload library per project with a tight inline file picker, real progress bar, and size limit warning.",
    ),
    "2026-05-22": (
        "Two big features: step copy/move + the for-each loop.",
        "Steps can be reordered, copied, or moved between flows. New 'for-each' lets one step repeat for every item in a list (e.g., check 100 product URLs with one config).",
    ),
    "2026-05-25": (
        "Drag-and-drop steps + nested loops (up to 4 levels) + split the 'check' and 'report' buttons.",
        "Replaced arrow buttons with industry-standard drag-and-drop. Loops can now nest (for each country, for each product, etc). Snapshot is now read-only.",
    ),
    "2026-05-27": (
        "Made the 'Run full check' button show live progress.",
        "A sticky banner at the top tells the user which phase is running (refreshing tokens, then checking URLs, then running each flow one by one).",
    ),
    "2026-05-29": (
        "Added a Compute step + live URL preview.",
        "Users can transform data between API calls (split text, slice arrays, add fields) without writing code. The URL field now shows the final URL as you type.",
    ),
    "2026-06-01": (
        "Added a pure Loop step, live loop progress, a 'concat arrays' transform, and the per-item version of it.",
        "Loop = repeat without a fake API call (~88 fewer calls per run). Live loop progress bar. 'Concat arrays' merges two lists. New: 'add a merged-list field to every item' — for each campaign, merge that campaign's countries + regions into one list and Loop it. Dogfooded on the real Logitech flow.",
    ),
}

# ---- Task ID -> (plain_task, plain_notes) ----
# Tone: everyday words. Allowed common tech: API, database, frontend, backend,
# URL, button, Slack, Postman, JSON, file, dashboard, login, token.
TASKS = {
    # ===== 2026-04-29 =====
    "1.1": ("Set up the backend server (the bit that runs the checks).",
            "Used Node.js + TypeScript. Server auto-restarts when I edit code."),
    "1.2": ("Set up the frontend web page (the bit users see).",
            "Used React + Vite. The web page talks to the backend through a proxy."),
    "1.3": ("Built the 'add a URL to watch' API.",
            "Users can add or remove URLs through simple API calls."),
    "1.4": ("Wrote code that actually visits a URL and checks if it works.",
            "Captures the HTTP status code (200 means OK, 404 means missing, etc)."),
    "1.5": ("Grouped URLs by health bucket.",
            "2xx = good, 3xx = redirect, 4xx = client error, 5xx = server broken, error = couldn't reach."),

    # ===== 2026-04-30 =====
    "1.6": ("Built the dashboard with colored count cards.",
            "Each card shows how many URLs are in each health bucket. Color-coded."),
    "1.7": ("Made the web page refresh by itself every 3 seconds.",
            "No need to hit F5 — new check results appear automatically."),
    "1.8": ("Saved data to a file safely.",
            "Used a 'write to temp file then rename' trick so the file is never half-written if the server crashes."),

    # ===== 2026-05-05 =====
    "2.1": ("Added support for multiple projects with a sidebar.",
            "Each project = one service being monitored (e.g. 'My API', 'Payment Service')."),
    "2.2": ("Added a per-project vault to store API keys.",
            "User picks the header name and prefix (like 'Bearer ...' or 'x-api-key: ...')."),
    "2.3": ("Made checks send the API key in the right header.",
            "Supports Bearer tokens, x-api-key, and Basic Auth."),
    "2.4": ("Let users pick how often each URL gets checked.",
            "Default: every 5 minutes. Range: 1 minute to 1 day."),
    "2.5": ("Tracked WHERE the time is going on slow requests.",
            "Splits the request time into 5 phases: DNS, TCP, TLS, server thinking, downloading."),
    "2.6": ("Let users add a free-text label to each URL.",
            "E.g. 'Login endpoint' or 'Get user profile'."),
    "2.7": ("Translated cryptic errors into plain English.",
            "Instead of 'ENOTFOUND', shows 'DNS lookup failed'. Instead of '401', shows 'wrong API key'."),
    "2.8": ("Added Slack alerts when something starts failing.",
            "Fires once when a URL goes from OK to broken (not every 5 minutes after)."),
    "2.9": ("Drew a colored bar showing where each request spends its time.",
            "5 colored segments per check — easy to spot 'oh, this one's slow because of TLS'."),
    "9.4": ("Wrote a doc explaining how Phase 1 connects to the Phase 2 plan.",
            "Mapped each Phase 2 requirement to the file in Phase 1 that already covers it."),

    # ===== 2026-05-06 =====
    "9.1": ("Demo: showed how API keys actually work (with and without).",
            "Hit the same URL with and without the key — got 401 vs 200."),
    "9.2": ("Demo: showed Basic Auth (username + password) too.",
            "Used a public test URL to prove the auth header really goes through."),

    # ===== 2026-05-07 =====
    "3.1": ("Built a proper popup dialog system.",
            "Replaced the ugly browser 'Are you sure?' boxes with nice in-app dialogs."),
    "3.2": ("Added toast notifications.",
            "Small floating messages in the corner (success / error). Auto-disappear after 3.5 seconds."),
    "3.3": ("Set up a design 'token' system in CSS.",
            "All spacing / fonts / corners / animation timings come from one place — easy to tweak consistently."),
    "3.4": ("Polished the sidebar.",
            "Project initials in colored circles + a small health dot per project."),
    "3.5": ("Added smooth animations everywhere.",
            "Cards lift on hover, dialogs slide up from the bottom — feels alive."),

    # ===== 2026-05-08 =====
    "4.1": ("Moved data from a plain JSON file to a real database (SQLite).",
            "Used Node's built-in SQLite — no extra installs needed."),
    "4.2": ("Designed the database tables.",
            "Tables: projects, API keys, URLs, check results. Linked properly with indexes for speed."),
    "4.3": ("Wrote a one-time migration: old JSON file → new database.",
            "Old file is backed up as .migrated.bak in case anything goes wrong."),
    "4.4": ("Started saving every single check result forever.",
            "Each check goes in the database with a timestamp."),
    "4.5": ("Added auto-cleanup so the database doesn't grow forever.",
            "Old check results (older than 7 days at first) get deleted hourly."),
    "5.1": ("Drew a tiny line chart (sparkline) showing the last 24h of speed per URL.",
            "Pure SVG, no chart library."),
    "5.2": ("Built a 24-hour status strip (commit-graph style — one box per check).",
            "(Later replaced by the Activity Timeline below.)"),
    "5.3": ("Added a 'fail rate' chip per URL with severity colors.",
            "Green if <1% fail, amber if 1–5%, red if more."),
    "5.4": ("Built the top KPI bar.",
            "4 big numbers per project: total URLs, % healthy, avg speed, # failing — plus a sparkline."),
    "5.5": ("Built the 'get check history for a URL' API.",
            "Returns all checks since a given timestamp."),
    "5.6": ("Built the 'get stats for a URL' API.",
            "Returns total checks, fail count, average speed, slowest speed."),
    "7.1": ("Added a 'Check All Now' button.",
            "Checks every URL in the project at once (up to 8 in parallel)."),
    "7.2": ("Generated a standalone HTML report after a full check.",
            "Saved to a reports folder — opens in any browser."),
    "7.3": ("Sent the result to Slack in a rich format.",
            "Block Kit message with KPIs and a list of what's failing."),
    "7.4": ("Attached the HTML report as a file to the Slack message.",
            "Used the official Slack file upload API."),
    "7.5": ("Showed a progress popup while the full check runs + a summary after.",
            "Spinner during, summary modal after."),

    # ===== 2026-05-11 =====
    "5.7": ("Built the Activity Timeline (unified history chart).",
            "Replaces the old status strip + sparkline with one cleaner view."),
    "6.1": ("Added GET, POST, PUT, PATCH method support.",
            "DELETE blocked on purpose (too dangerous to fire automatically)."),
    "6.2": ("Built the request body editor.",
            "Pick body type: JSON, form fields, or url-encoded."),
    "6.3": ("Added a custom headers tab.",
            "Simple key-value editor — sent on every check."),
    "6.4": ("Added a query parameters tab.",
            "Key-value editor — appended to the URL automatically."),
    "6.5": ("Built the assertions engine (4 starting types).",
            "Check that: status equals X / status in range / response faster than X ms / body contains text."),
    "6.6": ("Showed assertion results as pills on each URL card.",
            "Green check if all pass, red X if any fail."),
    "8.1": ("Added a search bar — filters URLs as you type.",
            "Matches URL, description, or method."),
    "8.2": ("Added numbered pagination (LeetCode style).",
            "Smart '...' when there are many pages."),
    "8.3": ("Added method filter chips with colors.",
            "GET = green, POST = orange, PUT = cyan, PATCH = purple."),
    "8.4": ("Added a '/' keyboard shortcut to jump to the search bar.",
            "Press '/' anywhere to focus the search box."),
    "8.5": ("Added a result count chip in the search bar.",
            "Turns blue when a filter is active."),
    "8.6": ("Made newly-added URLs show up at the top of the list.",
            "Sorted by row ID descending."),
    "9.3": ("Demo: showed POST with a JSON body really sending.",
            "Used the public httpbin test URL — confirmed body arrives correctly."),

    # ===== 2026-05-12 =====
    "4.6": ("Extended check history from 7 days to 365 days.",
            "Now you can look at a full year of trends."),
    "5.8": ("Added a time-range picker (24h / 7d / 30d / 90d / 1y / Custom).",
            "LinkedIn-style. All charts re-render based on the selection."),
    "5.9": ("Charts now auto-format their X-axis labels based on the range.",
            "Hours for 24h, days for a week, dates for a year, etc."),
    "5.10": ("Redesigned the time-range picker.",
             "Segmented pill with a sliding indicator — looks like Datadog or Linear."),
    "6.7": ("Added a 'Raw' body type with custom Content-Type.",
            "Send Text, XML, HTML, JS, YAML, or any custom type — just like Postman's Raw mode."),
    "12.1": ("Added skeleton loaders on first page load.",
             "Shimmering placeholder boxes show while the first data arrives — no more blank screen."),
    "12.2": ("Added a proper spinner inside async buttons.",
             "Instead of the button text changing to 'Checking…', a real spinner spins."),
    "12.3": ("Added icons + colored borders to toast notifications.",
             "Success / error / info each have their own SVG icon and color."),
    "12.4": ("Made failing project dots in the sidebar gently pulse.",
             "Pulls the eye to broken things without being annoying."),
    "12.5": ("URL cards now fade + slide in one after another.",
             "40ms stagger between cards — feels alive on first paint."),
    "12.6": ("Smoothed every status pill, KPI number, and chip transition.",
             "No more sudden color or number snaps."),
    "12.7": ("Added consistent focus rings everywhere for keyboard users.",
             "Tab through the app and you can always see what's focused."),
    "12.8": ("Respected the OS 'reduce motion' setting.",
             "Animations get toned down for users who prefer less motion."),

    # ===== 2026-05-19 (massive — Flows + Prereqs + UX hardening) =====
    "13.1": ("Designed the database tables for Flows.",
             "5 new tables: flows, flow_steps, flow_runs, step_results, variable_cache."),
    "13.2": ("Wrote the 'pull a value out of a JSON response' helper.",
             "Mini JSONPath: e.g. $.user.id grabs the user ID from the response."),
    "13.3": ("Wrote variable substitution.",
             "Anything like {{token}} in a URL/header/body gets replaced with the saved value before sending."),
    "13.4": ("Built the Flow runner — runs steps one after another.",
             "If any step fails, the whole flow stops and downstream steps are marked skipped."),
    "13.5": ("Added smart caching with a time-to-live.",
             "If the login token is still fresh, skip the login step — saves money on every run."),
    "13.6": ("Added per-step retries with exponential backoff.",
             "If a step fails due to a network blip, try again 1s later, then 2s, then 4s. Kills false alerts."),
    "13.7": ("Added a custom wait between steps.",
             "For async APIs: 'POST a job, wait 5 seconds, then poll for the result.'"),
    "13.8": ("Made the scheduler run whole flows on their interval.",
             "One timer per flow, not per step."),
    "13.9": ("Sent a Slack alert when a flow fails.",
             "Uses the same Slack settings as the per-URL alerts."),
    "13.10": ("Built 13 new API routes for managing flows.",
              "Create, edit, delete, reorder steps, trigger manually, view runs, view cached variables."),
    "13.11": ("Built the 'edit a flow' popup.",
              "Set name, run interval, and 'stop on failure' toggle."),
    "13.12": ("Built the 'edit a step' popup with 7 tabs.",
              "Basics / Params / Headers / Body / Assertions / Extract / Retry — full Postman parity."),
    "13.13": ("Step editor shows which variables are available from earlier steps.",
              "Self-documenting — you can see {{token}} is ready to use without flipping back."),
    "13.14": ("Built the Flow card on the dashboard.",
              "Click to expand the step list + see the result of the last run inline."),
    "13.15": ("Added a 'Run Now' button with a spinner + last-run timestamp.",
              "Trigger a flow on demand instead of waiting for the interval."),
    "13.16": ("Put the Flows section above standalone URLs in each project.",
              "Clean separation between flows and one-off URL checks."),
    "13.17": ("Wrapped Flows + URLs in proper visual section panels.",
              "Each section has its own header, count, and action buttons (Notion / Linear style)."),
    "13.18": ("Added GitHub-style tabs ('Standalone URLs' / 'Flows') with count badges.",
              "Each tab has its own search and filters — scales to many items."),
    "13.19": ("Made the active tab survive a page refresh (in the URL hash).",
              "Plus a red 'failing count' badge so you see breakage without switching tabs."),
    "13.20": ("Switching projects in the sidebar always opens the URLs tab.",
              "Avoids the confusion of 'I was on Flows in A, clicking B also opens Flows in B'."),
    "13.21": ("Tightened the vertical spacing between sections.",
              "Less wasted whitespace; better breathing room where it matters."),
    "13.22": ("'Check All' button now also re-runs all flows.",
              "One button audits everything in a project."),
    "13.23": ("HTML report now has a dedicated Flows section.",
              "URLs table + Flows table + per-step status badges + 4 KPIs."),
    "13.24": ("Slack message split into two tracks: URLs and Flows.",
              "Separate stats and separate 'failing' lists for each type."),
    "13.25": ("Sped up the list query with a single SQL join.",
              "Pulls each flow's last-run result in one trip — no N+1 query problem."),
    "13.26": ("KPI bar now understands flows.",
              "Label changes to 'Endpoints' and shows a breakdown — works fine with zero flows too."),
    "13.27": ("Audit result popup shows URL count vs Flow count.",
              "Clear separation in the result feedback."),
    "13.28": ("Flows tab gets its own mini KPI strip.",
              "Total / Healthy / Failing / Avg run time / Last run — at-a-glance flow health."),
    "13.29": ("Made the flow KPI strip bigger + added tooltips.",
              "Last-run cell shows the flow name as subtext. Every cell has a hover explanation."),
    "13.30": ("Fixed: flow KPI strip auto-refreshes after 'Run Now'.",
              "No more needing to reload the page."),
    "14.1": ("Designed 4 new database tables for Prerequisites.",
             "Plus extra columns on projects for the 'shared variable pool'."),
    "14.2": ("Built the create/read/update/delete code for prereq steps + runs + the project-wide variable pool.",
             "Mirrors the Flows store but variables live at the project level."),
    "14.3": ("Built the Prereq runner.",
             "Runs steps in order with retries, waits, caching. Captured values land in the project pool."),
    "14.4": ("Each URL check now auto-uses values from the project pool.",
             "Custom headers and body now resolve {{auth_token}} etc — login token stays warm."),
    "14.5": ("Flow runner merges the project pool with the flow's own variables.",
             "If both have the same name, the flow's value wins (more specific scope)."),
    "14.6": ("Scheduler now auto-runs the prereq chain before URL / flow checks.",
             "Login tokens stay fresh without anyone hitting 'refresh'."),
    "14.7": ("Built 7 new API routes for prereqs.",
             "CRUD prereq steps, run the chain, list runs, get vars, clear vars."),
    "14.8": ("Built the Prereqs panel that sits above the tabs.",
             "Collapsible. Shows status + 'Run Now' button. Visible on both tabs."),
    "14.9": ("Built the 'edit a prereq step' form.",
             "Shares the same 7-tab UX as the flow step editor."),
    "14.10": ("Step editor's variable hints now also list prereq-chain variables.",
              "So users discover {{auth_token}} without docs."),
    "14.11": ("Panel shows the live variable list with countdown timers + a 'Clear' button.",
              "See exactly what's cached and how long it's good for."),
    "14.12": ("Added per-project schedule controls (interval + enable/disable).",
              "Each project can have its own prereq refresh cadence."),
    "14.13": ("Smoke test: prereq captures token → URL check uses it → assertion passes.",
              "End-to-end proven with public test URLs."),
    "14.14": ("{{var}} substitution now works inside assertion config too.",
              "E.g. 'body should contain {{session_id}}' — fixes a brittleness gap."),
    "14.15": ("All three runners (URL / flow / prereq) now use the same variable list when checking assertions.",
              "Single source of truth — no surprises."),
    "14.16": ("Assertion UI hints that {{var}} works.",
              "Placeholder text + a tip line — users find the feature without reading docs."),
    "14.17": ("Smoke test: clear the variable pool → flow stays green because prereq auto-reruns.",
              "Proven across body, header, and query-param substitution."),
    "14.18": ("Split runners into 'kick off' (instant) + 'execute in the background'.",
              "New /run-async endpoints return a run ID immediately so the UI can show live progress."),
    "14.19": ("Frontend polls the run ID every 500ms and shows live per-step state.",
              "Pulsing 'running' pill, dashed 'queued' pill, filling progress bar."),
    "14.20": ("Progress bar now says 'Step 3 of 5 running…' with a completed count.",
              "No more opaque spinner."),
    "14.21": ("Added a 'force' flag — manual clicks bypass the smart cache.",
              "Human click = always fresh. Scheduler = still uses cache."),
    "14.22": ("All manual 'Run Now' buttons pass force=true.",
              "Button now does what it says — no confusing 'SKIPPED' message on click."),
    "14.23": ("Smoke test: scheduler skips when fresh, manual click always refreshes.",
              "Verified the variable pool value before and after to prove both paths."),
    "15.1": ("Active project survives a page refresh.",
             "Saved in localStorage — F5 keeps you where you were."),
    "15.2": ("Each project remembers your scroll position when you switch away.",
             "Comes back to the same spot when you switch back."),
    "15.3": ("Prereq panel auto-collapses 1.5 seconds after a run finishes.",
             "Expands during the run to show progress, then snaps back."),
    "15.4": ("Replaced browser confirm popups with a two-click inline confirm for step delete.",
             "Stays in-app — no jarring browser dialog inside our own popup."),
    "15.5": ("Browser tab title shows the active project + failing count.",
             "E.g. '(3 failing) My Project — Functional Monitor'."),
    "15.6": ("Sidebar shows a pulsing red 'failing count' badge per project.",
             "Spot broken projects at a glance without opening each one."),
    "15.7": ("Long URLs in step rows get clean truncation (… at the end).",
             "Stops them from pushing the 'edit' hint off the card."),

    # ===== 2026-05-20 =====
    "15.8": ("Backend now publishes each retry attempt + backoff phase in memory.",
             "Live retry visibility without writing every attempt to the database."),
    "15.9": ("'Get run details' API includes the live in-flight step.",
             "Same endpoint, same poll rate — frontend just gets richer data."),
    "15.10": ("Step pill switches to amber 'RETRY 2/3' during retries.",
              "Plus the whole row tints amber — easy to tell a retry from a fresh try."),
    "15.11": ("Progress bar says 'retry 2 of 3 (waiting before next try)' + shows the last status code.",
              "No more silent waits during backoff."),
    "15.12": ("Smoke test: hit a URL that returns 503, watched it retry 4 times live.",
              "Verified the attempt 1 → 2 → 3 → 4 transitions including the backoff waits."),

    # ===== 2026-05-21 =====
    "16.1": ("Created a database table + an on-disk folder for uploaded files.",
             "One row per file. Actual bytes live on disk (not in the database)."),
    "16.2": ("Centralized the 'where files live on disk' helper.",
             "One place that knows the uploads folder + the path for a given file ID."),
    "16.3": ("Built upload / download / delete API routes (10 MB cap).",
             "No multer dependency — raw bytes go straight to disk."),
    "16.4": ("Built database helpers: createUpload / getUpload / list per project / delete.",
             "Same pattern as everything else."),
    "16.5": ("Step body editor learned a 'binary' type.",
             "Pick a file from the project library. Empty field name = raw bytes. Set = multipart form."),
    "16.6": ("Built a shared BinaryBodyEditor component.",
             "Reused by the URL form + Flow step + Prereq step editors — write once, use everywhere."),
    "16.7": ("Added a Binary tab to all three body editors.",
             "Sits next to JSON / Raw / form-data — discoverable in the same place."),
    "16.8": ("Run Now on a Flow now auto-runs the prereq chain first (with force=true).",
             "Avoids spurious failures from stale login tokens."),
    "16.9": ("Renamed 'Audit' to 'Generate report' — it just snapshots current state now.",
             "Used to re-run everything; now defaults to fast snapshot. Re-check is opt-in."),
    "16.10": ("Smoke test: upload + list + download + delete round-trip via the API.",
              "Verified each step with raw API calls."),
    "16.11": ("Rebuilt the binary file picker to match Postman's style exactly.",
              "Compact 'Select File' button + inline filename + small X to clear — no oversized dropzone."),
    "16.12": ("Added a real upload progress bar.",
              "Live percentage as the file uploads — no more opaque 'Uploading…' text."),
    "16.13": ("Added a client-side 10 MB size check.",
              "Inline error shown before hitting the server — saves a wasted upload."),
    "16.14": ("Filename + size shown inline with a small X to clear.",
              "Matches Postman's binary body editor exactly."),
    "16.15": ("'Field name (optional)' is a single inline input.",
              "Empty = raw bytes. Set = multipart. Postman-style single-tab semantic."),
    "16.16": ("Inline left-bar error messages (no big banners).",
              "Postman-style subdued feedback."),
    "16.17": ("Project uploads library: tight collapsible rows.",
              "Thumbnail / extension / check mark if active / hover-only delete button."),
    "16.18": ("Flow card shows a 'Refreshing access tokens…' banner during prereq phase.",
              "Previously only a small pill — banner makes the wait self-explanatory."),
    "16.19": ("Audit button copy = 'Snapshot & report' + a tooltip explaining no re-checks happen.",
              "Button now reads as the action it actually performs."),
    "16.20": ("Fixed: clicking X on a selected file now deletes it from the project.",
              "Previously it stayed in the library — confusing."),
    "16.21": ("Replaced browser confirm for upload delete with the two-click inline confirm.",
              "Matches the step-delete pattern."),

    # ===== 2026-05-22 =====
    "16.22": ("Prereq banner now shows live step-N-of-M + completed count + retry chip + progress bar.",
              "Previously opaque — users couldn't tell if it was progressing or stuck."),
    "16.23": ("Flow card's Run Now now also shows the prereq progress UI in the panel above.",
              "User feedback: the full progress bar only appeared when triggered from the panel itself."),
    "17.1": ("Backend: copy-step and move-step between flows.",
             "Single safe database transaction — no half-finished state if it fails."),
    "17.2": ("New 'copy to flow' / 'move to flow' API routes.",
             "Validates: target flow exists, not the same flow, etc."),
    "17.3": ("Frontend API helpers for reorder / copy / move steps.",
             "Mirrors the existing reorder helper."),
    "17.4": ("Built a 'find {{variable}} references in a step' helper.",
             "Scans URL, body, headers, query for {{name}} tokens — reusable across flow + prereq steps."),
    "17.5": ("Added up/down arrow buttons on every step row.",
             "First step's 'up' is disabled. Last step's 'down' is disabled. Disabled mid-run."),
    "17.6": ("Hover-revealed 'Move' + 'Copy' buttons on flow step rows.",
             "Opens the move/copy popup. Only on flow steps (not prereq — by design)."),
    "17.7": ("Built the Move/Copy popup.",
             "Lists other flows in the project. New step always lands at position 1 of the target."),
    "17.8": ("Warn chip on rows with broken {{var}} references after reorder.",
             "Non-blocking — runtime still tries it and shows the real error."),
    "17.9": ("Smoke test: copy + move + reorder verified via the API.",
             "Positions correctly renumbered on both sides. Edge cases tested."),
    "17.10": ("Fixed: target flow auto-refreshes after Move/Copy.",
              "Previously needed a manual reload to see the moved step."),
    "17.11": ("Fixed: positions get renumbered when a step is deleted.",
              "Was: 1/2/3 → delete #2 → 1/3 (sparse). Now: → 1/2 (tight)."),
    "17.12": ("Scroll position now survives a full page reload.",
              "Saved to localStorage on tab close. Restored on next load. Was in-memory only before."),
    "17.13": ("Switching back to a project restores the last-viewed tab.",
              "If you were on 'Flows' in Project A, returning to A reopens 'Flows'. Not the default 'URLs'."),
    "18.1": ("Backend: ForEachConfig type + a forEach field on every step.",
             "Single-level only at first — one for-each per flow."),
    "18.2": ("Step results now carry iteration index + iteration count.",
             "Same count across every row of one iteration set."),
    "18.3": ("Extraction helper learned the [*] wildcard.",
             "$.data[*] returns the whole array. $.data[*].id returns the list of IDs."),
    "18.4": ("Variable substitution learned dotted paths.",
             "{{student.id}} walks into the object — bridges per-iteration items to template syntax."),
    "18.5": ("Database migrations for new columns (idempotent — safe to re-run).",
             "for_each_config_json on steps + iteration_index/_count on results."),
    "18.6": ("Backend validation: identifier names + 'only one for-each per flow' guard.",
             "Returns a clear 400 error on violation."),
    "18.7": ("Database read/write for the new iteration columns.",
             "Prereq runner passes nulls (prereqs never iterate)."),
    "18.8": ("Flow runner now forks into iterations.",
             "Resolves the array variable, caps at 100, loops with per-iteration vars, never stops the flow if one iteration fails."),
    "18.9": ("Live progress now publishes 'iteration X of N' between iterations.",
             "Mid-flight UI gets the count for free."),
    "18.10": ("Widened the variable map to support any type.",
              "Strings, arrays, objects — all stored properly. JSON-stringified only when written to disk."),
    "18.11": ("Frontend types mirror backend types (single source of truth).",
              "ForEachConfig + iteration fields aligned."),
    "18.12": ("New 'For each' tab in the step editor.",
              "Dropdown of array-typed variables + 'as' input + a disable button + a 'single-level only' note."),
    "18.13": ("Step header now shows a 'for each {{item}}' pill.",
              "Indigo tint — subtle but discoverable."),
    "18.14": ("Step row shows '(N) ✓ X / ✗ Y' summary + chevron to expand per-iteration details.",
              "Doesn't render 100 inline rows — uses a scrollable panel instead."),
    "18.15": ("Live progress label gets 'iteration X of N' when in a for-each step.",
              "Increments per iteration."),
    "18.16": ("Variable reference checker now recognises {{name.dotted.path}}.",
              "Plus it knows the loop-local item name so no false 'undefined' warnings."),
    "18.17": ("CSS for the new for-each pill / iteration summary / panel / warning banner.",
              "Visual weight matches the existing extract-row editor."),
    "18.18": ("Clean build: backend + frontend both compile with zero warnings.",
              "Bundle: 273 KB JS / 58 KB CSS."),

    # ===== 2026-05-25 =====
    "18.1.1": ("Frontend: drag-and-drop step reorder with optimistic UI + rollback on failure.",
               "Replaces the up/down arrows."),
    "18.1.2": ("Each step row now has a grip handle (6-dot icon).",
               "Drop-zone math splits the row by cursor Y vs midpoint (above / below)."),
    "18.1.3": ("Same drag-and-drop treatment in the Prereqs panel.",
               "Shared code path — same styles."),
    "18.1.4": ("CSS: replaced arrow buttons with a grip + drag visuals.",
               "2px accent line shows insertion point. Grab cursor → grabbing → not-allowed during run."),
    "18.1.5": ("Drop indicator hides for no-op hovers.",
               "Dragging row N over its own above-line doesn't flash a misleading line."),
    "18.1.6": ("Smoke test: API reorder verified on the for-each demo flow.",
               "Order swapped + restored cleanly."),
    "18.1.7": ("Clean build: zero warnings. Bundle 275 KB JS / 58 KB CSS.",
               "Drag-and-drop ships."),
    "18.2.1": ("Added the dnd-kit library (industry standard — Linear / Notion / Vercel use it).",
               "Bundle grew by ~17 KB gzip — worth it for proper accessibility and feel."),
    "18.2.2": ("Built a shared StepDragHandle component (grip icon + floating preview).",
               "SVG dots scale crisply at any screen density."),
    "18.2.3": ("Flow card step list now uses dnd-kit's DndContext + DragOverlay.",
               "Vertical sort + closest-center collision + 220ms eased drop animation."),
    "18.2.4": ("Same dnd-kit treatment in the Prereqs panel.",
               "Identical UX across both panels."),
    "18.2.5": ("Sensors for mouse + keyboard.",
               "Tab to grip, Space to grab, arrow keys to move, Space to drop, Esc to cancel. Full accessibility."),
    "18.2.6": ("Drop handler uses dnd-kit's arrayMove + optimistic swap + rollback on API failure.",
               "Same reorder API as before."),
    "18.2.7": ("CSS polish: focus ring on grip, 35% opacity while dragging, lifted shadow + scale + tilt on the preview.",
               "Production-grade visual polish."),
    "18.2.8": ("Removed the manual insertion-line indicator.",
               "dnd-kit's 'slide out of the way' animation makes the drop position obvious without it."),
    "18.2.9": ("Clean build: zero warnings. Bundle 327 KB JS / 59 KB CSS.",
               "Phase 1.18.2 ships."),
    "19.1": ("Types updated: for-each can now use dotted paths (e.g. student.subjects).",
             "Step result tracks iteration path for nested iteration."),
    "19.2": ("Database migrations: iteration_path + iteration_path_count columns.",
             "Old rows default to NULL = depth-1 (still readable)."),
    "19.3": ("Variable resolution now walks innermost-first.",
             "Inner loops shadow outer scopes — same behaviour as a programming language."),
    "19.4": ("Backend validation: for-each depth cannot exceed 4.",
             "Clear 400 error: 'for-each depth cannot exceed 4 (got N)'."),
    "19.5": ("Backend: dotted-path array names supported. Iteration path is saved + read.",
             "Identifier regex widened to accept dotted variant."),
    "19.6": ("Live progress now publishes the full path of iterations.",
             "Per-level cap (100) unchanged; total call cap (10,000) protects against 100^4 blow-ups."),
    "19.7": ("Flow runner rewritten: handles nested loops via implicit scope-stack.",
             "Walks contiguous for-each steps automatically — user never picks a parent loop."),
    "19.8": ("New recursive runner for nested for-each blocks.",
             "Per-iteration scope push/pop, direct-child recursion, all iterations continue at every level."),
    "19.9": ("Total-call-budget guard.",
             "If a run hits 10,000 total calls, it emits a 'Truncated' sentinel row and stops that branch."),
    "19.10": ("Frontend types updated to mirror backend.",
              "Single source of truth."),
    "19.11": ("Variable reference checker walks all earlier for-each steps.",
              "Nested {{student.id}} + {{subject.id}} + {{mark.id}} resolve without false warnings."),
    "19.12": ("Rewrote the For-each editor.",
              "Grouped dropdown (extracted vars vs loop items), 1–4 depth badge, dotted-path text input."),
    "19.13": ("Live warning banner: 'This step will run up to ~10,000 times per flow run.'",
              "Sets expectations before hitting the cap."),
    "19.14": ("Removed the old 'only one for-each' lock — multi-loop is now allowed.",
              "Dropped the warning banner that came with it."),
    "19.15": ("Loop pill gets depth color (1=teal, 2=violet, 3=amber, 4=rose).",
              "Live progress label shows the full path: 'iteration 3/10 → 7/12 → 2/8'."),
    "19.16": ("New IterationTree component — chevron-expandable per-level breadcrumb.",
              "Color stripe per depth, 16px indent, branch-level ok/fail aggregation."),
    "19.17": ("Frontend depth computation mirrors backend exactly.",
              "Same algorithm in 3 places — depth pill color stays in sync with the runner."),
    "19.18": ("CSS for depth pills + estimate banner + iteration tree.",
              "Visual weight matches existing Phase 1.18 styling."),
    "19.19": ("Clean build: 333 KB JS / 62 KB CSS.",
              "Nested for-each ships."),
    "19.1.1": ("Step results now save the resolved URL (after variable substitution).",
               "Each leaf in the iteration tree shows the actual fetched URL — not the template."),
    "19.1.2": ("Database migration for the new resolved_url column.",
               "Idempotent — picks up on next backend start."),
    "19.1.3": ("Database read/write code updated for the new column.",
               "Symmetric across flow + prereq tables (25-col / 21-col inserts)."),
    "19.1.4": ("Flow runner: success path saves the resolved URL.",
               "Sentinel rows (truncated / missing-var / cache-skip / upstream-failed) stay null since no real fetch happened."),
    "19.1.5": ("Prereq runner: same treatment.",
               "Symmetric with the flow runner."),
    "19.1.6": ("Frontend type updated.",
               "Single field add."),
    "19.1.7": ("Iteration tree now shows the resolved URL under each iteration row.",
               "Monospace + truncated + full URL on hover."),
    "19.1.8": ("CSS for the iteration-tree URL line.",
               "Scoped — doesn't bleed into other URL displays."),
    "19.1.9": ("Clean build + smoke test: 17/17 iteration rows now carry the resolved URL.",
               "Same bundle size as before."),
    "20.1": ("Snapshot/Report is now strictly read-only.",
             "Removed the 're-check first' option — per manager mandate, snapshot only reports what's currently known."),
    "20.2": ("Dropped the refresh query parameter from the audit API.",
             "Cleanly separates 'trigger a check' from 'generate a report'."),
    "20.3": ("New API: 'Check all URLs now' (up to 8 in parallel).",
             "Powers a new button — checks all URLs but not flows."),
    "20.4": ("New API: 'Run full check' (prereqs → URLs + flows).",
             "Prereqs first (so tokens are fresh), then URLs + flows in parallel. Continues even if prereqs fail."),
    "20.5": ("Frontend API helpers for the two new buttons.",
             "Typed result interfaces."),
    "20.6": ("Frontend wired up the two new actions with busy state + toast summary on completion.",
             "Disables the buttons during the run to prevent double-fire."),
    "20.7": ("Button 1: 'Check all now' added next to the HTTP method chips.",
             "Disabled when there are no URLs."),
    "20.8": ("Button 2: 'Run full check' added to the left of 'Snapshot & report'.",
             "Snapshot demoted to a secondary action — the new full-check button is the primary."),
    "20.9": ("CSS: method-chips row + check-all button.",
             "Chips bar + trigger button share one line."),
    "20.10": ("Clean build: 336 KB JS / 63 KB CSS.",
              "Phase 1.20 ships."),

    # ===== 2026-05-27 =====
    "20.1.1": ("Flow card can now 'attach' to a run that was kicked off by the parent.",
               "Polls the run ID without re-kicking — mirrors the existing prereq panel pattern."),
    "20.1.2": ("Refactored handleRun to accept options.",
               "Single polling path serves both self-triggered and parent-orchestrated runs."),
    "20.1.3": ("New orchestrator state in ProjectView drives the 'Run full check' flow.",
               "Runs prereqs → URLs in parallel → flows one at a time. Each phase has visible progress."),
    "20.1.4": ("New sticky banner at the top: phase-aware status text + spinner.",
               "Refreshing prerequisites → Checking URLs + running flows → Running flow X of Y: name → Full check complete."),
    "20.1.5": ("The matching flow card on the page lights up with live step-by-step progress.",
               "Exactly one card shows progress at a time (matched by flow ID)."),
    "20.1.6": ("Removed the old single-shot 'full check' code from App.tsx.",
               "ProjectView now drives everything through the orchestrator."),
    "20.1.7": ("CSS: sticky banner with blue→violet gradient + soft blur + fade-in.",
               "Keeps progress visible even while scrolling."),
    "20.1.8": ("Backend 'full check' API kept intact for scripts / cron.",
               "UI no longer uses it — but external callers still can."),
    "20.1.9": ("Clean build: 339 KB JS / 64 KB CSS.",
               "Phase 1.20.1 ships."),

    # ===== 2026-05-29 =====
    "21.1": ("Backend types: ComputeTransform (8 kinds), ComputeRow, ComputeConfig, StepType.",
             "Steps can now be 'http' or 'compute'. Existing steps default to 'http' — no break."),
    "21.2": ("Pure helper that applies a compute transform.",
             "Plus a URL preview helper that returns colored segments (literal / resolved / unresolved)."),
    "21.3": ("Database migrations: step_type column (default 'http') + compute_config_json (nullable).",
             "Existing rows untouched."),
    "21.4": ("Database read/write for the new compute columns on flow + prereq tables.",
             "Sentinel URL ('compute://step') keeps the existing NOT NULL constraint happy."),
    "21.5": ("Flow + Prereq runners handle Compute as a top-level branch.",
             "Within one Compute step, rows chain (each row can use the previous row's result)."),
    "21.6": ("New API: 'sample variables for this flow' — pulls the latest successful run.",
             "Powers the live URL preview panel in the step editor."),
    "21.7": ("Frontend types + API helpers for Compute + sample vars.",
             "Same shape as backend."),
    "21.8": ("New StepTypePicker (HTTP vs Compute) + new ComputeStepBody + per-kind compute row editor.",
             "8 transform kinds: splitTake, slice, replace, concat, mapAddField, etc. Chained rows in one step."),
    "21.9": ("Live URL preview panel under the URL input.",
             "Shows the final URL with sample variables filled in — updates live as you type."),
    "21.10": ("CSS for the new picker + compute editor + URL preview panel.",
              "Empty / loading / error / warn variants. Fits the existing visual language."),
    "21.11": ("Clean build: 353 KB JS / 67 KB CSS.",
              "Phase 1.21 ships."),

    # ===== 2026-06-01 =====
    "22.1": ("Backend: added 'loop' as a third step type (alongside http and compute).",
             "No new config shape needed — Loop reuses the existing for-each field."),
    "22.2": ("Backend store: 3-way switch in step create/update for the new Loop type.",
             "Loop steps use a sentinel URL ('loop://step') and need both 'iterate over' and 'as' fields."),
    "22.3": ("Backend runner: Loop iterations only push/pop scope (no HTTP call).",
             "After the loop finishes, emits ONE summary result row ('ran N iterations') — not N noisy rows."),
    "22.4": ("Frontend types + API helpers accept the new 'loop' type.",
             "Same shape as backend."),
    "22.5": ("Frontend: added a third amber Loop card to the step type picker.",
             "New LoopStepBody = description + the standalone for-each editor. Prereq form hides Loop."),
    "22.6": ("Step row swaps the method tag for a '🔁 LOOP' badge.",
             "Result chip says 'ran N iterations' instead of a meaningless latency number."),
    "22.7": ("Migrated the live Logitech flow via the API.",
             "Deleted the dummy httpbin scaffold steps + added real Loop steps. Cut ~88 wasted API calls per run."),
    "22.1.1": ("Frontend: Loop step now reads live progress from the inner running step.",
               "Pill flips from 'QUEUED' to '▶ ITER X/N' as soon as the loop body starts firing."),
    "22.1.2": ("New inline progress widget on Loop steps.",
               "Amber band with '🔁 iteration N of M' + right-aligned percentage + 4px gradient progress bar."),
    "22.1.3": ("CSS: ~50 lines for the loop-progress family.",
               "Verified live: forEachPath=[10,1] totals=[46,2] renders '▶ ITER 10/46' + 21% bar."),
    "22.2.1": ("Backend: added a 9th Compute transform — 'concat arrays' (merge two lists into one).",
               "Manager's idea: keep countries + regions in separate fields, then merge + Loop together."),
    "22.2.2": ("Backend: implemented the merge logic in the Compute engine.",
               "Looks up each source array by name, checks each is a list, returns one flat combined list. Missing names are skipped quietly."),
    "22.2.3": ("Frontend: added the new transform to the Compute editor.",
               "New 'Concat arrays (merge lists)' option in the kind dropdown. Editor shows numbered 'Source array N' inputs with add/remove buttons (supports any number of source lists, not just 2)."),
    "22.2.4": ("Built a demo flow end-to-end and confirmed it works.",
               "Flow: seed POST returns countries+regions → Compute merges them into geos → Loop fires 5 times → /US, /UK, /JP, /EMEA, /APAC each return 200 OK."),
    "22.2.5": ("Fixed a hidden bug user spotted in the URL preview panel.",
               "Stored arrays come back from the database as text. The preview was silently treating them as strings, so it never showed the per-iteration URLs — for ANY flow, not just the new demo. Endpoint now converts those strings back into arrays before returning."),
    "22.2.6": ("Fixed another hidden bug — dashboard iteration rows now show the actual URL fetched.",
               "Before: rows just said '#1 ✓ 200 2936ms'. Now they say '#1 ✓ 200 2936ms /anything/US/discover'. Data was already there in the database; the rendering just dropped it for single-loop flows (nested loops already showed it)."),
    "22.2.7": ("Made click-to-copy URL work in every project, not just Logitech.",
               "User noticed clicking a URL in the Logitech flow copies it but the same gesture didn't work elsewhere. Wasn't a project-specific gate — Logitech happened to use nested loops which used a different renderer that already had the copy feature. Single-loop flows used a plain span. Now both use the same copy-on-click button."),
    "22.3.1": ("Backend: 'merge lists per item' now actually works (the inner transform can see the item).",
               "When the Compute step runs 'for each campaign, add a merged list of its countries + regions', the merge step needs to read fields off the CURRENT campaign. Backend now hands the current campaign into the inner step so it can find 'countries' and 'regions' on it. Older inner transforms (split text, slice, etc.) are unaffected."),
    "22.3.2": ("Frontend: 'merge lists' is now selectable as the inner step + has its own input boxes.",
               "Two small fixes: (a) the inner-step dropdown was filtering out 'merge lists' — un-filtered it; (b) when you pick it, the editor now shows the same 'Source list 1 / Source list 2 / + Add another list' inputs as the top-level version, plus a short explainer note."),
    "22.3.3": ("Rewrote the live Logitech flow to use the new per-item merge (real dogfood).",
               "Step 2 used to derive the locale via 'split text take part 0'; now Step 2 derives 'campaign.geos = countries + regions' per campaign. Step 4 iterates campaign.geos with the new URL format /campaign/{{documentId}}/{{locale}}/{{geo}}/discover. Result: 66/69 calls green; remaining 3 are real backend signal (zh-Hant locale 403s) — exactly what monitoring is for."),

    # ===== Pending (no date — kept for completeness) =====
    "7.6": ("Pending: get a Slack bot token (xoxb-) and paste it in Settings.",
            "Enables the Slack file upload feature."),
    "9.5": ("Pending: schedule a manager meeting to present the plan and get approval.",
            "User to schedule."),
}


# ---- Date -> ordered list of task IDs for that date ----
# (Authoritative — does not depend on CSV parsing, since CSV has bare commas
# embedded in some Notes fields.)
DATE_TASKS = {
    "2026-04-29": ["1.1", "1.2", "1.3", "1.4", "1.5"],
    "2026-04-30": ["1.6", "1.7", "1.8"],
    "2026-05-05": ["2.1", "2.2", "2.3", "2.4", "2.5", "2.6", "2.7", "2.8", "2.9", "9.4"],
    "2026-05-06": ["9.1", "9.2"],
    "2026-05-07": ["3.1", "3.2", "3.3", "3.4", "3.5"],
    "2026-05-08": ["4.1", "4.2", "4.3", "4.4", "4.5",
                   "5.1", "5.2", "5.3", "5.4", "5.5", "5.6",
                   "7.1", "7.2", "7.3", "7.4", "7.5"],
    "2026-05-11": ["5.7", "6.1", "6.2", "6.3", "6.4", "6.5", "6.6",
                   "8.1", "8.2", "8.3", "8.4", "8.5", "8.6", "9.3"],
    "2026-05-12": ["4.6", "5.8", "5.9", "5.10", "6.7",
                   "12.1", "12.2", "12.3", "12.4", "12.5", "12.6", "12.7", "12.8"],
    "2026-05-19": [
        "13.1", "13.2", "13.3", "13.4", "13.5", "13.6", "13.7", "13.8", "13.9", "13.10",
        "13.11", "13.12", "13.13", "13.14", "13.15", "13.16", "13.17", "13.18", "13.19", "13.20",
        "13.21", "13.22", "13.23", "13.24", "13.25", "13.26", "13.27", "13.28", "13.29", "13.30",
        "14.1", "14.2", "14.3", "14.4", "14.5", "14.6", "14.7", "14.8", "14.9", "14.10",
        "14.11", "14.12", "14.13", "14.14", "14.15", "14.16", "14.17", "14.18", "14.19", "14.20",
        "14.21", "14.22", "14.23",
        "15.1", "15.2", "15.3", "15.4", "15.5", "15.6", "15.7",
    ],
    "2026-05-20": ["15.8", "15.9", "15.10", "15.11", "15.12"],
    "2026-05-21": ["16.1", "16.2", "16.3", "16.4", "16.5", "16.6", "16.7", "16.8", "16.9", "16.10",
                   "16.11", "16.12", "16.13", "16.14", "16.15", "16.16", "16.17", "16.18", "16.19", "16.20", "16.21"],
    "2026-05-22": ["16.22", "16.23",
                   "17.1", "17.2", "17.3", "17.4", "17.5", "17.6", "17.7", "17.8", "17.9",
                   "17.10", "17.11", "17.12", "17.13",
                   "18.1", "18.2", "18.3", "18.4", "18.5", "18.6", "18.7", "18.8", "18.9", "18.10",
                   "18.11", "18.12", "18.13", "18.14", "18.15", "18.16", "18.17", "18.18"],
    "2026-05-25": [
        "18.1.1", "18.1.2", "18.1.3", "18.1.4", "18.1.5", "18.1.6", "18.1.7",
        "18.2.1", "18.2.2", "18.2.3", "18.2.4", "18.2.5", "18.2.6", "18.2.7", "18.2.8", "18.2.9",
        "19.1", "19.2", "19.3", "19.4", "19.5", "19.6", "19.7", "19.8", "19.9",
        "19.10", "19.11", "19.12", "19.13", "19.14", "19.15", "19.16", "19.17", "19.18", "19.19",
        "19.1.1", "19.1.2", "19.1.3", "19.1.4", "19.1.5", "19.1.6", "19.1.7", "19.1.8", "19.1.9",
        "20.1", "20.2", "20.3", "20.4", "20.5", "20.6", "20.7", "20.8", "20.9", "20.10",
    ],
    "2026-05-27": ["20.1.1", "20.1.2", "20.1.3", "20.1.4", "20.1.5", "20.1.6", "20.1.7", "20.1.8", "20.1.9"],
    "2026-05-29": ["21.1", "21.2", "21.3", "21.4", "21.5", "21.6", "21.7", "21.8", "21.9", "21.10", "21.11"],
    "2026-06-01": ["22.1", "22.2", "22.3", "22.4", "22.5", "22.6", "22.7",
                   "22.1.1", "22.1.2", "22.1.3",
                   "22.2.1", "22.2.2", "22.2.3", "22.2.4", "22.2.5", "22.2.6", "22.2.7",
                   "22.3.1", "22.3.2", "22.3.3"],
}

PENDING_IDS = ["7.6", "9.5"]


def build_workbook():
    # Sort dates chronologically
    sorted_dates = sorted(DATE_TASKS.keys())
    by_date = DATE_TASKS
    pending = PENDING_IDS

    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Log"

    # ---- Styles ----
    thin = Side(border_style="thin", color="D5D5D5")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="334155")  # slate
    header_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    day_font = Font(name="Calibri", size=12, bold=True, color="78350F")  # amber-900
    day_fill = PatternFill("solid", fgColor="FEF3C7")  # amber-100
    day_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    summary_font = Font(name="Calibri", size=10, italic=True, color="92400E")  # amber-700
    summary_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    id_font = Font(name="Calibri", size=10, color="64748B")  # slate-500
    task_font = Font(name="Calibri", size=11, color="0F172A")  # slate-900
    notes_font = Font(name="Calibri", size=10, color="475569")  # slate-600
    task_align = Alignment(horizontal="left", vertical="top", wrap_text=True)

    pending_font = Font(name="Calibri", size=11, bold=True, color="92400E")
    pending_fill = PatternFill("solid", fgColor="FEF3C7")

    # ---- Header row ----
    ws.cell(row=1, column=1, value="Date / #").font = header_font
    ws.cell(row=1, column=2, value="What I Built").font = header_font
    ws.cell(row=1, column=3, value="Why / Details").font = header_font
    for col in range(1, 4):
        c = ws.cell(row=1, column=col)
        c.fill = header_fill
        c.alignment = header_align
        c.border = border
    ws.row_dimensions[1].height = 22

    current_row = 2

    # ---- Per-day blocks ----
    for date in sorted_dates:
        items = by_date[date]
        summary = DAY_SUMMARIES.get(date, ("", ""))

        # Day header (merged): "YYYY-MM-DD  -  N tasks"
        day_label = f"{date}   -   {len(items)} task{'s' if len(items) != 1 else ''}"

        # Row 1 of day block: merged label
        ws.cell(row=current_row, column=1, value=day_label).font = day_font
        ws.cell(row=current_row, column=1).fill = day_fill
        ws.cell(row=current_row, column=1).alignment = day_align
        ws.cell(row=current_row, column=1).border = border
        # Merge col 1 across cols 1 to 3
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
        for col in range(1, 4):
            ws.cell(row=current_row, column=col).fill = day_fill
            ws.cell(row=current_row, column=col).border = border
        ws.row_dimensions[current_row].height = 24
        current_row += 1

        # Row 2 of day block: 2-line summary
        summary_text = f"{summary[0]}\n{summary[1]}" if summary[0] else ""
        if summary_text:
            ws.cell(row=current_row, column=1, value=summary_text).font = summary_font
            ws.cell(row=current_row, column=1).fill = day_fill
            ws.cell(row=current_row, column=1).alignment = summary_align
            ws.cell(row=current_row, column=1).border = border
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
            for col in range(1, 4):
                ws.cell(row=current_row, column=col).fill = day_fill
                ws.cell(row=current_row, column=col).border = border
            ws.row_dimensions[current_row].height = 42
            current_row += 1

        # Individual task rows
        for task_id in items:
            plain = TASKS.get(task_id, (f"(missing entry for {task_id})", ""))
            plain_task, plain_notes = plain

            c1 = ws.cell(row=current_row, column=1, value=task_id)
            c1.font = id_font
            c1.alignment = Alignment(horizontal="left", vertical="top")
            c1.border = border

            c2 = ws.cell(row=current_row, column=2, value=plain_task)
            c2.font = task_font
            c2.alignment = task_align
            c2.border = border

            c3 = ws.cell(row=current_row, column=3, value=plain_notes)
            c3.font = notes_font
            c3.alignment = task_align
            c3.border = border

            # Auto height based on text length (rough)
            longest = max(len(plain_task or ""), len(plain_notes or ""))
            if longest > 120:
                ws.row_dimensions[current_row].height = 48
            elif longest > 80:
                ws.row_dimensions[current_row].height = 34
            else:
                ws.row_dimensions[current_row].height = 22

            current_row += 1

    # ---- Pending block ----
    if pending:
        ws.cell(row=current_row, column=1, value="Pending (no date yet)").font = pending_font
        ws.cell(row=current_row, column=1).fill = pending_fill
        ws.cell(row=current_row, column=1).alignment = day_align
        ws.cell(row=current_row, column=1).border = border
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
        for col in range(1, 4):
            ws.cell(row=current_row, column=col).fill = pending_fill
            ws.cell(row=current_row, column=col).border = border
        ws.row_dimensions[current_row].height = 24
        current_row += 1

        for task_id in pending:
            plain = TASKS.get(task_id, (f"(missing entry for {task_id})", ""))
            plain_task, plain_notes = plain

            c1 = ws.cell(row=current_row, column=1, value=task_id)
            c1.font = id_font
            c1.alignment = Alignment(horizontal="left", vertical="top")
            c1.border = border

            c2 = ws.cell(row=current_row, column=2, value=plain_task)
            c2.font = task_font
            c2.alignment = task_align
            c2.border = border

            c3 = ws.cell(row=current_row, column=3, value=plain_notes)
            c3.font = notes_font
            c3.alignment = task_align
            c3.border = border

            ws.row_dimensions[current_row].height = 28
            current_row += 1

    # ---- Column widths ----
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 70
    ws.column_dimensions["C"].width = 75

    # ---- Freeze header ----
    ws.freeze_panes = "A2"

    wb.save(OUT)
    print(f"Wrote {OUT}")
    print(f"Days covered: {len(sorted_dates)}, total tasks: {sum(len(v) for v in by_date.values())}, pending: {len(pending)}")


if __name__ == "__main__":
    build_workbook()
