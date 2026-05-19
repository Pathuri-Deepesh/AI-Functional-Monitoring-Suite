"""Reverse sync: read project-tracker.xlsx → write project-tracker.csv + PROGRESS.md."""
import csv
import sys
from pathlib import Path
from datetime import date
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
XLSX_PATH = ROOT / "project-tracker.xlsx"
CSV_PATH = ROOT / "project-tracker.csv"
MD_PATH = ROOT / "PROGRESS.md"

if not XLSX_PATH.exists():
    print(f"missing: {XLSX_PATH}", file=sys.stderr)
    sys.exit(1)

wb = load_workbook(XLSX_PATH, data_only=True)
ws = wb.active

raw_rows = []
for r in ws.iter_rows(values_only=True):
    if r and any(cell not in (None, "") for cell in r):
        raw_rows.append([("" if v is None else str(v)).strip() for v in r])

if not raw_rows:
    print("xlsx is empty", file=sys.stderr)
    sys.exit(1)

header = raw_rows[0]
data = raw_rows[1:]

# ---------- write CSV ----------
with CSV_PATH.open("w", encoding="utf-8", newline="") as f:
    writer = csv.writer(f, lineterminator="\n")
    writer.writerow(header)
    for row in data:
        # pad/truncate to header length
        padded = (row + [""] * len(header))[: len(header)]
        writer.writerow(padded)
print(f"wrote: {CSV_PATH} ({len(data)} rows)")

# ---------- write PROGRESS.md ----------
# Group rows by Phase
from collections import OrderedDict

phases = OrderedDict()
for row in data:
    if len(row) < 5:
        continue
    num, phase, task, status, done_on = row[0], row[1], row[2], row[3], row[4]
    notes = row[5] if len(row) > 5 else ""
    phases.setdefault(phase, []).append({
        "num": num, "task": task, "status": status, "done_on": done_on, "notes": notes,
    })

today = date.today().isoformat()

lines = []
lines.append("# Project Progress Tracker — AI-Powered Functional Monitoring Suite")
lines.append("")
lines.append("**Owner:** Deepesh P · **Company:** Logitech · **Started:** 2026-04-29")
lines.append("")
lines.append("> Source of truth: `project-tracker.xlsx`. This file is auto-generated from it.")
lines.append("> Edit XLSX in Excel, save, then ask Claude to sync.")
lines.append("")

for phase_name, items in phases.items():
    pending = [i for i in items if i["status"].lower() != "done"]
    suffix = " ✅ Complete" if not pending else f" ({len(pending)} pending)"
    lines.append(f"## {phase_name}{suffix}")
    lines.append("")
    for item in items:
        ticked = "x" if item["status"].lower() == "done" else " "
        date_part = f"*{item['done_on']}*" if item["done_on"] else "*date: ____________*"
        lines.append(f"- [{ticked}] **{item['num']}** {item['task']} — {date_part}")
    lines.append("")

lines.append("---")
lines.append("")
lines.append(f"*Last sync from XLSX: {today}*")
lines.append("")

MD_PATH.write_text("\n".join(lines), encoding="utf-8")
print(f"wrote: {MD_PATH}")
