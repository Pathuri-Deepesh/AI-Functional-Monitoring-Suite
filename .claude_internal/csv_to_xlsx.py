"""One-shot script: convert project-tracker.csv → project-tracker.xlsx with formatting."""
import csv
import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
CSV_PATH = ROOT / "project-tracker.csv"
XLSX_PATH = ROOT / "project-tracker.xlsx"

if not CSV_PATH.exists():
    print(f"missing: {CSV_PATH}", file=sys.stderr)
    sys.exit(1)

with CSV_PATH.open("r", encoding="utf-8") as f:
    rows = list(csv.reader(f))

wb = Workbook()
ws = wb.active
ws.title = "Tracker"

# Styles
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", fgColor="1F2937")
header_align = Alignment(horizontal="left", vertical="center")

phase_fills = {
    "Phase 1": "DBEAFE",          # light blue
    "Phase 1.5": "DCFCE7",        # light green
    "Phase 1.6": "FEF3C7",        # light amber
    "Phase 1.7": "EDE9FE",        # light purple
    "Phase 1.8": "FCE7F3",        # light pink
    "Phase 1.9": "DBEAFE",
    "Phase 1.10": "DCFCE7",
    "Phase 1.11": "FEF3C7",
    "Demo": "F3F4F6",
    "Phase 2": "FEE2E2",          # light red — upcoming
    "Documentation": "F3F4F6",
}

status_colors = {"Done": "10B981", "In Progress": "F59E0B", "Pending": "9CA3AF"}

thin_border = Border(
    left=Side(style="thin", color="E5E7EB"),
    right=Side(style="thin", color="E5E7EB"),
    top=Side(style="thin", color="E5E7EB"),
    bottom=Side(style="thin", color="E5E7EB"),
)

for r, row in enumerate(rows, start=1):
    for c, value in enumerate(row, start=1):
        cell = ws.cell(row=r, column=c, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        if r == 1:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
        else:
            phase = row[1] if len(row) > 1 else ""
            for prefix, fill in phase_fills.items():
                if phase.startswith(prefix):
                    cell.fill = PatternFill("solid", fgColor=fill)
                    break
            if c == 4:  # Status column
                status = (value or "").strip()
                color = status_colors.get(status)
                if color:
                    cell.font = Font(bold=True, color=color)

ws.freeze_panes = "A2"
ws.auto_filter.ref = ws.dimensions

# Reasonable column widths
widths = {1: 6, 2: 20, 3: 70, 4: 14, 5: 14, 6: 50}
for col_idx, width in widths.items():
    ws.column_dimensions[get_column_letter(col_idx)].width = width

wb.save(XLSX_PATH)
print(f"wrote: {XLSX_PATH}")
print(f"rows:  {len(rows) - 1} tasks")
